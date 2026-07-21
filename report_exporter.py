"""
报表导出模块 (ReportExporter)
支持将汇算结果、物料库、BOM 数据导出为 Excel (.xlsx) 或 CSV 格式。
仅依赖 openpyxl + csv，无需 pandas。
"""

import csv
import json
import os
from datetime import datetime
from typing import Optional, List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import BOMDatabase, MaterialDatabase
from ref_designator import format_designators


# ---------- 样式常量 ----------

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, col_count: int):
    """为工作表的第一行应用表头样式。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """自动调整列宽。"""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = min(length + 2, max_width)
        ws.column_dimensions[col_letter].width = max_len


def _write_csv(rows: List[Dict], output_path: str):
    """将字典列表写入 CSV 文件。"""
    if not rows:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("")
        return
    headers = list(rows[0].keys())
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_styled_excel(
    data: List[Dict],
    sheet_name: str,
    output_path: str,
    title: str = "",
):
    """将字典列表写入带样式的 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        return

    headers = list(data[0].keys())

    if title:
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for row_dict in data:
        ws.append([
            str(v) if v is not None else ""
            for v in row_dict.values()
        ])

    _auto_width(ws)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


class ReportExporter:
    """报表导出器。
    bom_db:      BOMDatabase（汇算、BOM数据）
    material_db: MaterialDatabase（物料、分类、制造商、系统配置）
    """

    def __init__(self, bom_db, material_db=None):
        self.db = bom_db
        self.material_db = material_db

    # ------------------------------------------------------------------
    # 汇算结果报表
    # ------------------------------------------------------------------

    def export_calculation(
        self,
        task_id: int,
        output_path: str,
        fmt: str = "xlsx",
    ) -> str:
        """
        导出汇算结果报表。

        包含三个工作表：汇算汇总、汇算明细（按来源展开）、汇算参数。
        返回输出文件的完整路径。
        """
        with self.db.get_connection() as conn:
            task = conn.execute(
                "SELECT * FROM calculation_tasks WHERE task_id = ?", (task_id,)
            ).fetchone()
            if not task:
                raise ValueError(f"汇算任务不存在: {task_id}")
            if task["status"] != "Completed":
                raise ValueError(f"汇算任务未完成，当前状态: {task['status']}")

            boms = conn.execute(
                """SELECT cb.bom_id, cb.order_quantity,
                          bh.board_name, bh.version
                   FROM calculation_boms cb
                   JOIN bom_headers bh ON cb.bom_id = bh.bom_id
                   WHERE cb.task_id = ?""",
                (task_id,),
            ).fetchall()

        # Cross-DB query for items with material details
        with self.db.cross_db_connection() as conn:
            items = conn.execute(
                """SELECT ci.*, m.mpn, m.description, m.footprint,
                          m.lifecycle_status, c.name AS category_name,
                          mf.name AS manufacturer_name
                   FROM calculation_items ci
                   JOIN mat.materials m ON ci.part_number = m.part_number
                   LEFT JOIN mat.categories c ON m.category_id = c.id
                   LEFT JOIN mat.manufacturers mf ON m.manufacturer_id = mf.id
                   WHERE ci.task_id = ?
                   ORDER BY c.code_prefix, ci.part_number""",
                (task_id,),
            ).fetchall()

        if fmt == "csv":
            return self._export_calculation_csv(items, boms, output_path)
        return self._export_calculation_xlsx(task, boms, items, output_path)

    def _export_calculation_xlsx(self, task, boms, items, output_path: str) -> str:
        wb = Workbook()

        # --- 工作表一：汇算汇总 ---
        ws1 = wb.active
        ws1.title = "汇算汇总"

        # Build BOM name list for dynamic columns
        bom_names = [f"{b['board_name']} {b['version']}" for b in boms]
        bom_keys = [(b['board_name'], b['version']) for b in boms]

        fixed_headers = [
            "物料编码", "分类", "描述", "封装",
            "制造商型号", "制造商",
            "理论需求量", "建议采购量",
            "实际备料数量", "备注",
        ]
        headers1 = fixed_headers + bom_names
        ws1.append(headers1)
        _style_header(ws1, len(headers1))

        for r in items:
            # Parse source_details to build a lookup by (board_name, version)
            src_map = {}
            try:
                for src in json.loads(r["source_details"] or "[]"):
                    key = (src.get("board_name", ""), src.get("version", ""))
                    src_map[key] = src.get("subtotal", 0)
            except (json.JSONDecodeError, TypeError):
                pass

            row_data = [
                r["part_number"],
                r["category_name"],
                r["description"],
                r["footprint"],
                r["mpn"],
                r["manufacturer_name"],
                r["theoretical_qty"],
                r["suggested_qty"],
                "",  # 实际备料数量 (空)
                "",  # 备注 (空)
            ]
            # Append per-BOM subtotal columns
            for bk in bom_keys:
                row_data.append(src_map.get(bk, ""))

            ws1.append(row_data)

            if r["lifecycle_status"] in ("NRND", "EOL"):
                for col in range(1, len(headers1) + 1):
                    ws1.cell(row=ws1.max_row, column=col).fill = _WARNING_FILL

        _auto_width(ws1)

        # --- 工作表二：汇算明细 ---
        ws2 = wb.create_sheet("汇算明细")
        headers2 = [
            "物料编码", "制造商型号", "描述",
            "来源BOM", "版本", "单板用量", "生产数量", "小计", "位号明细",
        ]
        ws2.append(headers2)
        _style_header(ws2, len(headers2))

        for r in items:
            sources = []
            try:
                sources = json.loads(r["source_details"])
            except (json.JSONDecodeError, TypeError):
                pass

            for src in sources:
                ws2.append([
                    r["part_number"], r["mpn"], r["description"],
                    src.get("board_name", ""),
                    src.get("version", ""),
                    src.get("unit_qty", 0),
                    src.get("order_qty", 0),
                    src.get("subtotal", 0),
                    src.get("ref_designators", ""),
                ])

        _auto_width(ws2)

        # --- 工作表三：汇算参数 ---
        ws3 = wb.create_sheet("汇算参数")
        ws3.append(["参数", "值"])
        _style_header(ws3, 2)

        ws3.append(["任务ID", task["task_id"]])
        ws3.append(["创建时间", task["created_at"]])
        ws3.append(["完成时间", task["completed_at"]])
        ws3.append(["执行耗时(ms)", task["duration_ms"]])
        ws3.append(["物料种类数", len(items)])
        ws3.append(["", ""])
        ws3.append(["包含BOM", ""])

        for b in boms:
            ws3.append([
                f"  {b['board_name']} {b['version']}",
                f"生产数量: {b['order_quantity']}",
            ])

        _auto_width(ws3)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)

        with self.db.transaction() as conn:
            conn.execute(
                "UPDATE calculation_tasks SET result_file_path = ? WHERE task_id = ?",
                (output_path, task["task_id"]),
            )

        return output_path

    def _export_calculation_csv(self, items, boms, output_path: str) -> str:
        bom_keys = [(b['board_name'], b['version']) for b in boms]
        bom_names = [f"{b['board_name']} {b['version']}" for b in boms]
        rows = []
        for r in items:
            src_map = {}
            try:
                for src in json.loads(r["source_details"] or "[]"):
                    key = (src.get("board_name", ""), src.get("version", ""))
                    src_map[key] = src.get("subtotal", 0)
            except (json.JSONDecodeError, TypeError):
                pass

            row = {
                "物料编码": r["part_number"],
                "分类": r["category_name"],
                "描述": r["description"],
                "封装": r["footprint"],
                "制造商型号": r["mpn"],
                "制造商": r["manufacturer_name"],
                "理论需求量": r["theoretical_qty"],
                "建议采购量": r["suggested_qty"],
                "实际备料数量": "",
                "备注": "",
            }
            for name, bk in zip(bom_names, bom_keys):
                row[name] = src_map.get(bk, "")
            rows.append(row)
        _write_csv(rows, output_path)
        return output_path

    # ------------------------------------------------------------------
    # 物料库导出
    # ------------------------------------------------------------------

    def export_materials(
        self,
        output_path: str,
        category: Optional[str] = None,
        status: Optional[str] = None,
        fmt: str = "xlsx",
    ) -> str:
        conditions = []
        params: list = []
        if category:
            conditions.append("c.code_prefix = ?")
            params.append(category)
        if status:
            conditions.append("m.lifecycle_status = ?")
            params.append(status)
        where = " AND ".join(conditions) if conditions else "1=1"

        with self.material_db.get_connection() as conn:
            rows = conn.execute(
                f"""SELECT m.part_number, mf.name AS manufacturer_name, m.mpn,
                           m.description, c.name AS category_name,
                           m.value, m.unit, m.footprint, m.lifecycle_status,
                           m.datasheet_url, m.default_loss_rate, m.moq, m.spq,
                           m.stock_qty, m.created_at
                    FROM materials m
                    LEFT JOIN categories c ON m.category_id = c.id
                    LEFT JOIN manufacturers mf ON m.manufacturer_id = mf.id
                    WHERE {where}
                    ORDER BY c.code_prefix, m.part_number""",
                params,
            ).fetchall()

        data = [
            {
                "物料编码": r["part_number"],
                "制造商": r["manufacturer_name"],
                "制造商型号": r["mpn"],
                "描述": r["description"],
                "分类": r["category_name"],
                "数值": r["value"],
                "单位": r["unit"],
                "封装": r["footprint"],
                "生命周期": r["lifecycle_status"],
                "规格书": r["datasheet_url"],
                "损耗率": r["default_loss_rate"],
                "MOQ": r["moq"],
                "SPQ": r["spq"],
                "库存": r["stock_qty"],
                "创建时间": r["created_at"],
            }
            for r in rows
        ]

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        if fmt == "csv":
            _write_csv(data, output_path)
        else:
            _write_styled_excel(data, "物料库", output_path)

        return output_path

    # ------------------------------------------------------------------
    # BOM 导出
    # ------------------------------------------------------------------

    def export_bom(
        self,
        bom_id: int,
        output_path: str,
        fmt: str = "xlsx",
    ) -> str:
        with self.db.get_connection() as conn:
            header = conn.execute(
                "SELECT * FROM bom_headers WHERE bom_id = ?", (bom_id,)
            ).fetchone()
            if not header:
                raise ValueError(f"BOM 不存在: {bom_id}")

        # Cross-DB query for items with material details
        with self.db.cross_db_connection() as conn:
            items = conn.execute(
                """SELECT bi.part_number, bi.quantity, bi.reference_designators,
                          m.mpn, m.description, m.footprint, m.lifecycle_status,
                          c.name AS category_name, mf.name AS manufacturer_name
                   FROM bom_items bi
                   JOIN mat.materials m ON bi.part_number = m.part_number
                   LEFT JOIN mat.categories c ON m.category_id = c.id
                   LEFT JOIN mat.manufacturers mf ON m.manufacturer_id = mf.id
                   WHERE bi.bom_id = ?
                   ORDER BY bi.part_number""",
                (bom_id,),
            ).fetchall()

        data = [
            {
                "物料编码": r["part_number"],
                "分类": r["category_name"],
                "描述": r["description"],
                "位号": r["reference_designators"] or "",
                "封装": r["footprint"],
                "数量": r["quantity"],
                "制造商型号": r["mpn"],
                "制造商": r["manufacturer_name"],
            }
            for r in items
        ]

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        if fmt == "csv":
            _write_csv(data, output_path)
        else:
            _write_styled_excel(
                data, "BOM", output_path,
                title=f"{header['board_name']} {header['version']}",
            )

        return output_path

    # ------------------------------------------------------------------
    # BOM 对比报表导出
    # ------------------------------------------------------------------

    def export_bom_diff(
        self,
        diff_result: Dict,
        bom_a_name: str,
        bom_b_name: str,
        output_path: str,
    ) -> str:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "新增物料"
        ws1.append(["物料编码", "MPN", "描述", "数量", "位号"])
        _style_header(ws1, 5)
        for item in diff_result.get("added", []):
            ws1.append([
                item.get("part_number"), item.get("mpn"),
                item.get("description"), item.get("quantity"),
                item.get("reference_designators", ""),
            ])
        _auto_width(ws1)

        ws2 = wb.create_sheet("删除物料")
        ws2.append(["物料编码", "MPN", "描述", "数量", "位号"])
        _style_header(ws2, 5)
        for item in diff_result.get("removed", []):
            ws2.append([
                item.get("part_number"), item.get("mpn"),
                item.get("description"), item.get("quantity"),
                item.get("reference_designators", ""),
            ])
        _auto_width(ws2)

        ws3 = wb.create_sheet("数量变更")
        ws3.append(["物料编码", "MPN", "描述", f"{bom_a_name}数量", f"{bom_b_name}数量", "变化量"])
        _style_header(ws3, 6)
        for item in diff_result.get("changed", []):
            ws3.append([
                item.get("part_number"), item.get("mpn"),
                item.get("description"),
                item.get("qty_a"), item.get("qty_b"), item.get("diff"),
            ])
        _auto_width(ws3)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        return output_path
